from app.celery_app import celery_app
from app.core.job_manager_file import (
    update_job_status,
    save_job_data,
)
import csv
import io
import json

try:
    import pandas as pd
except ImportError:  # pragma: no cover - runtime fallback
    pd = None


@celery_app.task(bind=True)
def process_file(self, job_id: str, content: bytes, filename: str):
    try:
        update_job_status(job_id, "processing", progress=10)

        if filename.endswith(".csv"):
            if pd is not None:
                df = pd.read_csv(io.BytesIO(content))
                data = df.to_dict(orient="list")
                columns = list(df.columns)
                rows = len(df)
            else:
                text = content.decode("utf-8-sig")
                rows = list(csv.DictReader(io.StringIO(text)))
                if not rows:
                    columns = []
                    data = {}
                else:
                    columns = list(rows[0].keys())
                    data = {col: [row.get(col, "") for row in rows] for col in columns}
                rows = len(rows)
        elif filename.endswith((".xlsx", ".xls")):
            try:
                from openpyxl import load_workbook
            except ImportError as exc:
                raise RuntimeError("Excel parsing requires openpyxl") from exc

            workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
            sheet = workbook.active
            rows = list(sheet.iter_rows(values_only=True))
            if not rows:
                columns = []
                data = {}
            else:
                headers = [str(cell) if cell is not None else "" for cell in rows[0]]
                columns = headers
                data = {
                    header: [row[idx] if idx < len(row) else "" for row in rows[1:]]
                    for idx, header in enumerate(headers)
                }
                rows = len(rows) - 1
        else:
            raise ValueError("Unsupported file format")

        update_job_status(job_id, "processing", progress=50)

        save_job_data(
            job_id,
            {
                "filename": filename,
                "data": data,
                "columns": columns,
                "rows": rows,
            },
        )

        update_job_status(job_id, "completed", progress=100)

    except Exception as exc:
        update_job_status(job_id, "failed", progress=0)
        raise exc
